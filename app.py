import base64
import os
import json
import time
import streamlit as st
import pandas as pd
from datetime import datetime
from mistralai import Mistral, JSONSchema, ResponseFormat
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Initialize Mistral client
api_key = os.environ.get("MISTRAL_API_KEY")
if not api_key:
    st.error("MISTRAL_API_KEY environment variable not set")
    st.stop()

client = Mistral(api_key=api_key)

# Create output folder if it doesn't exist
output_folder = "extracted_data"
Path(output_folder).mkdir(exist_ok=True)

# Master Facility List
FACILITY_LIST = [
    "Alliance Health at Marina Bay, 2 Seaport, Quincy, MA",
    "Alliance Health at West Acres, 804 Pleasant St, Brockton, MA",
    "Sherrill House, 135 S Huntington Ave, Jamaica Plain, MA",
    "Alliance Health at Maples, 90 Taunton St, Wrentham, MA 02097",
    "Oak Knoll, 9 Ambetter Dr, Framingham, MA",
    "Sippican Rehab & Healthcare, 15 Mill St, Marion, MA",
    "Alliance Health at Braintree, 175 Grove St, Braintree, MA",
    "Harrington House Rehab & Healthcare, 160 Main St, Walpole, MA",
    "Bethany Healthcare Rest Home, 97 Bethany Rd, Framingham, MA",
    "Alliance Health at Marie Esther, 720 Boston, Marlborough, MA",
    "Shrewsbury Nursing & Rehab Center, 40 Julio Dr, Shrewsbury, MA 01545",
    "Alliance Health at Doolittle Unit 1, 16 Bird St, Foxboro, MA",
    "CareOne at Concord, 57 Old Rd to 9 Acre Corner, Concord, MA 01742",
    "The Commons at Lincoln, 3 Harvest Cir, Lincoln, MA",
    "Rivercrest Nursing and Wellness, 100 Newbury Ct, Concord, MA",
    "Brookhaven at Lexington Independent Living, 1010 Waltham St, Lexington, MA",
    "Woburn Rehabilitation & Nursing Center, 18 Frances St, Woburn, MA 01801",
    "CareOne at Wilmington, 750 Woburn St, Wilmington, MA 01887",
    "CareOne at Lexington, 178 Lowell St, Lexington, MA 02420",
    "Winchester Rehabilitation and Nursing Center, 223 Swanton St, Winchester, MA 01890",
    "Aberjona Rehabilitation & Nursing Center, 184 Swanton St, Winchester, MA 01890",
    "The Commons in Lincoln, 1 Harvest Cir, Lincoln, MA 01773",
    "CareOne at Essex Park, 265 Essex St, Beverly, MA 01915",
    "CareOne at Peabody, 199 Andover St, Peabody, MA 01960"
]

# Fixed Excel headers
EXCEL_HEADERS = [
    "Phleb", "Date", "No of Patient", "Patient ID", "patient bod", 
    "Name of Patient", "Patient Birthdate", "Facility Information", 
    "Patients ICD Code", "From", "To", "Miles", "To_2", "Miles_2", 
    "To_3", "Miles_3", "To_4", "Miles_4", "To_5", "Miles_5", 
    "Total Miles", "Insurance Company", "Mem ID", "Group Mem ID"
]

def encode_file(file_content):
    """Encode file content to base64"""
    return base64.b64encode(file_content).decode('utf-8')

def extract_icd_codes_only(icd_string):
    """Extract only ICD codes from string, removing descriptions"""
    if not icd_string or icd_string.strip() == "":
        return ""
    
    import re
    # Pattern to match ICD codes (letter followed by numbers, optional dot and more characters)
    # Examples: I50.9, E11.9, T81.49XD, Z23, D64.9, L03.113
    icd_pattern = r'[A-Z]\d{2,3}(?:\.\d{1,4})?(?:[A-Z]{1,2})?'
    
    codes = re.findall(icd_pattern, icd_string.upper())
    
    # Remove duplicates while preserving order
    seen = set()
    unique_codes = []
    for code in codes:
        if code not in seen:
            seen.add(code)
            unique_codes.append(code)
    
    return ", ".join(unique_codes)

def match_facility(extracted_facility):
    """Match extracted facility name with master list and return matched facility"""
    if not extracted_facility or extracted_facility.strip() == "":
        return ""
    
    extracted_lower = extracted_facility.lower()
    
    # Remove everything after RM:, RM , Room:, Room  for cleaner matching
    for separator in [" rm:", " rm ", " room:", " room "]:
        if separator in extracted_lower:
            idx = extracted_lower.find(separator)
            extracted_lower = extracted_lower[:idx].strip()
            break
    
    # Remove common unit suffixes (ROCKPORT, JAMESPORT, EASTPORT, etc.)
    unit_suffixes = ["rockport", "jamesport", "eastport", "westport", "northport", "southport", 
                     "unit 1", "unit 2", "unit 3", "unit1", "unit2", "unit3"]
    for suffix in unit_suffixes:
        if suffix in extracted_lower:
            extracted_lower = extracted_lower.replace(suffix, "").strip()
    
    # First pass: Match by facility name (before comma in master list)
    for facility in FACILITY_LIST:
        facility_lower = facility.lower()
        facility_parts = [part.strip() for part in facility_lower.split(',')]
        facility_name = facility_parts[0]  # Get just the name part
        
        # Check if facility name is in extracted text
        if facility_name in extracted_lower:
            return facility
        
        # Also check if extracted text contains the facility name
        if extracted_lower in facility_name or facility_name in extracted_lower:
            return facility
    
    # Second pass: Match by key words (Alliance Health at Maples, CareOne at Concord, etc.)
    for facility in FACILITY_LIST:
        facility_lower = facility.lower()
        facility_name = facility_lower.split(',')[0].strip()
        
        # Split into words and check for significant matches
        facility_words = facility_name.split()
        extracted_words = extracted_lower.split()
        
        # Count matching words
        match_count = 0
        for word in facility_words:
            if len(word) > 2 and word in extracted_lower:
                match_count += 1
        
        # If most words match (at least 60%), consider it a match
        if len(facility_words) > 0 and match_count >= len(facility_words) * 0.6:
            return facility
    
    # No match found - return empty string
    return ""

def call_with_retry(api_func, max_retries=5, initial_delay=5):
    """Call API function with exponential backoff retry on rate limit errors"""
    delay = initial_delay
    for attempt in range(max_retries):
        try:
            return api_func()
        except Exception as e:
            error_str = str(e)
            # Check if it's a rate limit error (429)
            if "429" in error_str or "rate_limit" in error_str.lower() or "rate limit" in error_str.lower():
                if attempt < max_retries - 1:
                    st.warning(f"⏳ Rate limit hit. Waiting {delay} seconds before retry ({attempt + 1}/{max_retries})...")
                    time.sleep(delay)
                    delay *= 2  # Exponential backoff
                else:
                    raise Exception(f"Rate limit exceeded after {max_retries} retries. Please try again later or use a smaller PDF.")
            else:
                raise e

def split_pdf_to_images(pdf_content):
    """Convert PDF pages to individual images for batch processing"""
    import fitz  # PyMuPDF
    from io import BytesIO
    
    pdf_document = fitz.open(stream=pdf_content, filetype="pdf")
    page_images = []
    
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        # Render page to image (150 DPI for good quality without being too large)
        mat = fitz.Matrix(150/72, 150/72)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        page_images.append(base64.b64encode(img_bytes).decode('utf-8'))
    
    pdf_document.close()
    return page_images

def process_pdf(pdf_file, progress_callback=None):
    """Process PDF using Mistral OCR API with batch processing for large files"""
    try:
        # Update progress
        if progress_callback:
            progress_callback(5, "Reading PDF file...")
        
        # Read the file content
        pdf_content = pdf_file.read()
        pdf_size_mb = len(pdf_content) / (1024 * 1024)
        
        # Check if PDF is large (> 5MB or many pages)
        # For large PDFs, use page-by-page processing
        try:
            import fitz
            pdf_doc = fitz.open(stream=pdf_content, filetype="pdf")
            num_pages = len(pdf_doc)
            pdf_doc.close()
        except ImportError:
            # If PyMuPDF not available, estimate pages from size
            num_pages = max(1, int(pdf_size_mb * 10))  # Rough estimate
        
        if progress_callback:
            progress_callback(10, f"Detected {num_pages} pages...")
        
        all_text = ""
        
        # For large PDFs (>30 pages), process in batches
        if num_pages > 30:
            try:
                import fitz
                page_images = split_pdf_to_images(pdf_content)
                
                # Process in batches of 10 pages
                batch_size = 10
                total_batches = (len(page_images) + batch_size - 1) // batch_size
                
                for batch_idx in range(total_batches):
                    start_idx = batch_idx * batch_size
                    end_idx = min(start_idx + batch_size, len(page_images))
                    batch_images = page_images[start_idx:end_idx]
                    
                    progress_pct = 10 + int((batch_idx / total_batches) * 35)
                    if progress_callback:
                        progress_callback(progress_pct, f"Processing pages {start_idx + 1}-{end_idx} of {num_pages}...")
                    
                    # Process each page in the batch
                    for i, img_base64 in enumerate(batch_images):
                        page_num = start_idx + i + 1
                        
                        def ocr_call():
                            return client.ocr.process(
                                document={
                                    "type": "image_url",
                                    "image_url": f"data:image/png;base64,{img_base64}"
                                },
                                model="mistral-ocr-latest",
                                include_image_base64=False,
                            )
                        
                        ocr_response = call_with_retry(ocr_call)
                        
                        for page in ocr_response.pages:
                            all_text += f"\n--- Page {page_num} ---\n"
                            all_text += page.markdown + "\n"
                        
                        # Small delay between pages to avoid rate limits
                        time.sleep(0.5)
                    
                    # Longer delay between batches
                    if batch_idx < total_batches - 1:
                        time.sleep(2)
                
            except ImportError:
                st.error("For large PDFs (>30 pages), please install PyMuPDF: pip install pymupdf")
                return None
        else:
            # For smaller PDFs, use the original single-request method with retry
            if progress_callback:
                progress_callback(15, "Running OCR on document...")
            
            base64_file = base64.b64encode(pdf_content).decode('utf-8')
            
            def ocr_call():
                return client.ocr.process(
                    document={
                        "type": "document_url",
                        "document_url": f"data:application/pdf;base64,{base64_file}"
                    },
                    model="mistral-ocr-latest",
                    include_image_base64=False,
                )
            
            ocr_response = call_with_retry(ocr_call)
            
            if progress_callback:
                progress_callback(40, "Extracting text from pages...")
            
            for page in ocr_response.pages:
                all_text += page.markdown + "\n"
        
        # Update progress
        if progress_callback:
            progress_callback(50, "Analyzing document with AI...")
        
        # Step 2: Use chat completion to extract structured data
        extraction_prompt = """You are a medical document data extraction specialist. This PDF may contain multiple patient records and dropsheet pages.

DOCUMENT STRUCTURE:
- A "DROPSHEET" is a handwritten page with nurse names, patient lists, or signature sheets - it marks the START of a new section
- After each dropsheet, there are multiple patient report pages

YOUR TASK:
Extract ALL patient records from the document and return them as a JSON array. For each record, identify if it's a DROPSHEET marker or a PATIENT record.

RECORD TYPES:
1. "type": "dropsheet" - When you detect a handwritten dropsheet/signature page (set all other fields to empty "")
2. "type": "patient" - For actual patient data records

FIELDS TO EXTRACT FOR PATIENT RECORDS:

MANDATORY FIELDS (Must be extracted):
- "Date" - The date of service/visit
- "Name of Patient" - Full name of the patient
- "Facility Information" - Complete facility details including name, room number, full address, and phone number

CONDITIONALLY MANDATORY FIELDS (Extract if present):
- "Patient Birthdate" - Patient's date of birth
- "Patients ICD Code" - Medical diagnosis code (e.g., T81.49XD, Z23, etc.)
- "Insurance Company" - Name of insurance provider (look for "Insurance:", "Ins:", "Payer:", "Insurance Company:")
- "Mem ID" - Member ID number (look for "Member ID:", "Sub/Member No.:", "Subscriber ID:", "Member No:", "ID:", "Policy No:") - This is a NUMBER/ID, NOT an address!
- "Group Mem ID" - Group Member ID (look for "Group ID:", "Group No:", "Group Member ID:")

IMPORTANT FIELD DISTINCTIONS:
- "Ins Address:" or "Insurance Address:" is the ADDRESS of the insurance company - DO NOT put this in Mem ID!
- "Mem ID" should ONLY contain the member/subscriber ID NUMBER (alphanumeric like "3RW9K87UR58", "7RA2TG1JN99")
- Do NOT confuse address fields with ID fields

FIELDS TO LEAVE BLANK (Always ""):
- Phleb, No of Patient, Patient ID, patient bod, From, To, Miles, To_2, Miles_2, To_3, Miles_3, To_4, Miles_4, To_5, Miles_5, Total Miles

EXTRACTION RULES:
- Copy values EXACTLY as they appear
- Return an ARRAY of records in the order they appear in the document
- Insert a dropsheet record when you detect handwritten pages with nurse names/signatures
- NEVER guess or fabricate data

Return a JSON array like: [{"type": "dropsheet", ...}, {"type": "patient", ...}, {"type": "patient", ...}, {"type": "dropsheet", ...}, ...]"""
        
        def chat_call():
            return client.chat.complete(
                model="mistral-large-latest",
                messages=[
                    {
                        "role": "user",
                        "content": f"{extraction_prompt}\n\n--- DOCUMENT CONTENT ---\n{all_text}\n--- END OF DOCUMENT ---"
                    }
                ],
                response_format=ResponseFormat(
                    type="json_schema",
                    json_schema=JSONSchema(
                        name="response_schema",
                        schema_definition={
                            "type": "object",
                            "properties": {
                                "records": {
                                    "type": "array",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "type": {"type": "string", "enum": ["dropsheet", "patient"]},
                                            "Phleb": {"type": "string", "default": ""},
                                            "Date": {"type": "string"},
                                            "No of Patient": {"type": "string", "default": ""},
                                            "Patient ID": {"type": "string", "default": ""},
                                            "patient bod": {"type": "string", "default": ""},
                                            "Name of Patient": {"type": "string"},
                                            "Patient Birthdate": {"type": "string"},
                                            "Facility Information": {"type": "string"},
                                            "Patients ICD Code": {"type": "string"},
                                            "From": {"type": "string", "default": ""},
                                            "To": {"type": "string", "default": ""},
                                            "Miles": {"type": "string", "default": ""},
                                            "To_2": {"type": "string", "default": ""},
                                            "Miles_2": {"type": "string", "default": ""},
                                            "To_3": {"type": "string", "default": ""},
                                            "Miles_3": {"type": "string", "default": ""},
                                            "To_4": {"type": "string", "default": ""},
                                            "Miles_4": {"type": "string", "default": ""},
                                            "To_5": {"type": "string", "default": ""},
                                            "Miles_5": {"type": "string", "default": ""},
                                            "Total Miles": {"type": "string", "default": ""},
                                            "Insurance Company": {"type": "string"},
                                            "Mem ID": {"type": "string"},
                                            "Group Mem ID": {"type": "string"}
                                        },
                                        "required": ["type"]
                                    }
                                }
                            },
                            "required": ["records"]
                        },
                    ),
                ),
            )
        
        chat_response = call_with_retry(chat_call)
        
        # Update progress
        if progress_callback:
            progress_callback(85, "Parsing extracted data...")
        
        # Parse the JSON response
        extracted_data = json.loads(chat_response.choices[0].message.content)
        
        # Update progress
        if progress_callback:
            progress_callback(95, "Finalizing...")
        
        return extracted_data.get("records", [])
    
    except Exception as e:
        st.error(f"Error processing PDF: {str(e)}")
        return None

def save_to_excel(data, filename):
    """Save extracted data to Excel file with fixed headers, blank rows for dropsheets, and conditional highlighting"""
    try:
        filepath = os.path.join(output_folder, filename)
        
        # Build rows with blank rows for dropsheets
        rows = []
        
        # Track values within each section (reset on dropsheet)
        section_first_date = None
        section_last_facility = None  # Track last shown facility to show only on change
        
        if isinstance(data, list):
            for record in data:
                if isinstance(record, dict):
                    record_type = record.get("type", "patient")
                    
                    if record_type == "dropsheet":
                        # Add a blank row for dropsheet separator
                        blank_row = {header: "" for header in EXCEL_HEADERS}
                        rows.append(blank_row)
                        # Reset section tracking for new section
                        section_first_date = None
                        section_last_facility = None
                    else:
                        # Add patient data row
                        row_data = {header: record.get(header, "") for header in EXCEL_HEADERS}
                        
                        # Get current values
                        current_date = row_data.get("Date", "").strip()
                        current_facility_raw = row_data.get("Facility Information", "").strip()
                        
                        # Match facility with master list
                        matched_facility = match_facility(current_facility_raw)
                        row_data["Facility Information"] = matched_facility
                        
                        # Extract only ICD codes without descriptions
                        icd_raw = row_data.get("Patients ICD Code", "")
                        row_data["Patients ICD Code"] = extract_icd_codes_only(icd_raw)
                        
                        # Normalize date for comparison (lowercase, remove extra spaces)
                        current_date_normalized = " ".join(current_date.lower().split())
                        
                        # For facility comparison, use the matched facility name (before comma)
                        if matched_facility:
                            facility_name_for_compare = matched_facility.split(',')[0].strip().lower()
                        else:
                            facility_name_for_compare = ""
                        
                        # Handle Date - only show if different from first in section
                        if section_first_date is None:
                            # First record in section - keep the date and store it
                            section_first_date = current_date_normalized
                        elif current_date_normalized == section_first_date:
                            # Same as first in section - make blank
                            row_data["Date"] = ""
                        else:
                            # Different date - keep it (but don't update section_first_date)
                            pass
                        
                        # Handle Facility Information - only show if different from last shown facility
                        if section_last_facility is None:
                            # First record in section - keep the facility and store it
                            section_last_facility = facility_name_for_compare
                        elif facility_name_for_compare == section_last_facility:
                            # Same facility as last shown - make blank
                            row_data["Facility Information"] = ""
                        else:
                            # Different facility - keep it and update last shown facility
                            section_last_facility = facility_name_for_compare
                        
                        rows.append(row_data)
        else:
            # Single record as dictionary
            row_data = {header: data.get(header, "") for header in EXCEL_HEADERS}
            rows.append(row_data)
        
        # Create DataFrame
        df = pd.DataFrame(rows, columns=EXCEL_HEADERS)
        
        # Save to Excel
        df.to_excel(filepath, index=False, engine='openpyxl')
        
        # Apply conditional highlighting for Insurance fields
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Red fill for highlighting missing fields
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Find column indices (1-based for openpyxl)
        header_row = [cell.value for cell in ws[1]]
        insurance_col = header_row.index("Insurance Company") + 1 if "Insurance Company" in header_row else None
        mem_id_col = header_row.index("Mem ID") + 1 if "Mem ID" in header_row else None
        group_mem_id_col = header_row.index("Group Mem ID") + 1 if "Group Mem ID" in header_row else None
        
        # Apply highlighting rules for each data row (starting from row 2)
        for row in range(2, ws.max_row + 1):
            # Check if this is a blank row (dropsheet separator) - skip highlighting
            is_blank_row = all(ws.cell(row=row, column=col).value in [None, ""] for col in range(1, len(header_row) + 1))
            if is_blank_row:
                continue
            
            insurance_value = ws.cell(row=row, column=insurance_col).value if insurance_col else None
            mem_id_value = ws.cell(row=row, column=mem_id_col).value if mem_id_col else None
            group_mem_id_value = ws.cell(row=row, column=group_mem_id_col).value if group_mem_id_col else None
            
            # Check if each field is present (not empty)
            insurance_present = insurance_value and str(insurance_value).strip() != ""
            mem_id_present = mem_id_value and str(mem_id_value).strip() != ""
            group_mem_id_present = group_mem_id_value and str(group_mem_id_value).strip() != ""
            
            # Highlight each missing field in red individually
            if not insurance_present:
                ws.cell(row=row, column=insurance_col).fill = red_fill
            
            if not mem_id_present:
                ws.cell(row=row, column=mem_id_col).fill = red_fill
            
            if not group_mem_id_present:
                ws.cell(row=row, column=group_mem_id_col).fill = red_fill
        
        # Save the workbook with highlighting
        wb.save(filepath)
        wb.close()
        
        return filepath
    except Exception as e:
        st.error(f"Error saving to Excel: {str(e)}")
        return None

# Streamlit UI
st.set_page_config(page_title="PDF OCR Processor", layout="centered")

# Custom CSS for centered layout
st.markdown("""
<style>
    .main-title {
        text-align: center;
        padding: 2rem 0;
    }
    .upload-section {
        display: flex;
        flex-direction: column;
        align-items: center;
        padding: 2rem;
        border: 2px dashed #ccc;
        border-radius: 10px;
        margin: 2rem 0;
    }
    .stTabs [data-baseweb="tab-list"] {
        justify-content: center;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("<h1 class='main-title'>📄 PDF OCR Data Extractor</h1>", unsafe_allow_html=True)

# Initialize session state for storing results
if 'result' not in st.session_state:
    st.session_state.result = None
if 'filepath' not in st.session_state:
    st.session_state.filepath = None
if 'filename' not in st.session_state:
    st.session_state.filename = None

# Centered file upload section
st.markdown("---")
uploaded_file = st.file_uploader("📁 Upload a PDF file", type="pdf", key="pdf_uploader")

if uploaded_file:
    st.success(f"✓ File uploaded: {uploaded_file.name}")
    
    # Centered process button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🔄 Process PDF", key="process_btn", use_container_width=True):
            # Create progress bar container
            progress_container = st.empty()
            start_time = time.time()
            
            def update_progress(percent, status):
                """Update progress bar with percentage and estimated time remaining"""
                elapsed = time.time() - start_time
                if percent > 0:
                    estimated_total = elapsed / (percent / 100)
                    remaining = estimated_total - elapsed
                    
                    if remaining < 60:
                        time_str = f"{int(remaining)}s remaining"
                    else:
                        mins = int(remaining // 60)
                        secs = int(remaining % 60)
                        time_str = f"{mins}m {secs}s remaining"
                else:
                    time_str = "Calculating..."
                
                progress_container.progress(
                    percent / 100,
                    text=f"⏳ {percent}% complete - {status} ({time_str})"
                )
            
            # Process the PDF with progress tracking
            result = process_pdf(uploaded_file, update_progress)
            
            if result:
                st.session_state.result = result
                
                # Generate filename with timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{uploaded_file.name.replace('.pdf', '')}_{timestamp}.xlsx"
                st.session_state.filename = filename
                
                # Save to Excel
                update_progress(98, "Saving to Excel...")
                filepath = save_to_excel(result, filename)
                st.session_state.filepath = filepath
                
                # Show 100% complete briefly
                elapsed = time.time() - start_time
                progress_container.progress(1.0, text=f"✅ 100% complete - Done! (Total: {int(elapsed)}s)")
                time.sleep(1)
                
                # Clear progress and show success
                progress_container.empty()
                st.success("✓ PDF processed successfully!")

# Display results in tabs if data exists
if st.session_state.result and st.session_state.filepath:
    st.markdown("---")
    
    # Create two tabs
    tab1, tab2 = st.tabs(["📥 Download Excel", "📋 JSON Data"])
    
    with tab1:
        st.subheader("Download Extracted Data")
        st.write("Your extracted data is ready for download.")
        
        # Centered download button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            with open(st.session_state.filepath, 'rb') as f:
                excel_content = f.read()
            
            st.download_button(
                label="📥 Download Excel File",
                data=excel_content,
                file_name=st.session_state.filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        st.info(f"File saved to: {st.session_state.filepath}")
    
    with tab2:
        st.subheader("Extracted JSON Data")
        st.json(st.session_state.result)
